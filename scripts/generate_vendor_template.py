"""
generate_vendor_template.py
---------------------------
업체에게 보내줄 양식 엑셀(.xlsx)을 자동 생성한다. (대시보드 sky/blue 톤)

생성 위치: data/vendor_template.xlsx
시트 구성:
  - 안내      : 사용 방법, 컬럼 설명
  - 일일평가  : 매일 입력하는 평가 데이터 (헤더 + 예시 + 빈 입력행)
  - 에러로그  : 에러 발생 시마다 1행 추가

build_dashboard_json.py 가 인식하는 컬럼명·시트명을 그대로 사용한다.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "data" / "vendor_template.xlsx"

# ── 팔레트 (대시보드와 동일 톤) ───────────────────────────────
INK       = "0F2E54"   # deep navy (타이틀/텍스트)
NAVY      = "1E4A7A"   # 헤더 배경
SKY       = "2E89D6"   # 액센트
SKY_SOFT  = "EAF2FB"   # 섹션/예시 틴트
ZEBRA     = "F5F9FE"   # 교차 행
LINE      = "C9DAEE"   # 옅은 경계선
WHITE     = "FFFFFF"

FONT = "맑은 고딕"
BANNER_FONT   = Font(name=FONT, size=18, bold=True, color=WHITE)
BANNER_SUB    = Font(name=FONT, size=10, bold=True, color="BFD8F2")
HEADER_FONT   = Font(name=FONT, size=11, bold=True, color=WHITE)
SECTION_FONT  = Font(name=FONT, size=12, bold=True, color=INK)
BODY_FONT     = Font(name=FONT, size=10, color="2B3A4F")
NOTE_FONT     = Font(name=FONT, size=10, color="3D4147")
SUB_FONT      = Font(name=FONT, size=9, italic=True, color="6B7C93")
EXAMPLE_FONT  = Font(name=FONT, size=10, italic=True, color="8092A8")

BANNER_FILL  = PatternFill("solid", fgColor=INK)
HEADER_FILL  = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=SKY_SOFT)
EXAMPLE_FILL = PatternFill("solid", fgColor=SKY_SOFT)
ZEBRA_FILL   = PatternFill("solid", fgColor=ZEBRA)
WHITE_FILL   = PatternFill("solid", fgColor=WHITE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_T = Alignment(horizontal="left", vertical="top", wrap_text=True)

_thin = Side(border_style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HDR_BORDER = Border(left=_thin, right=_thin, top=_thin,
                    bottom=Side(border_style="medium", color=SKY))
SECTION_BORDER = Border(bottom=Side(border_style="thin", color=SKY))


# ── 데이터 시트 공통 빌더 ──────────────────────────────────────
def _style_header(ws: Worksheet, cols: list[str], widths: list[int]):
    for i, (name, width) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = HDR_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.sheet_view.showGridLines = False


def _example_rows(ws: Worksheet, start: int, rows: list[list], n_cols: int, height: int):
    for r, vals in enumerate(rows, start=start):
        for i in range(1, n_cols + 1):
            c = ws.cell(row=r, column=i, value=vals[i - 1] if i - 1 < len(vals) else "")
            c.fill = EXAMPLE_FILL
            c.font = EXAMPLE_FONT
            c.alignment = LEFT_T
            c.border = BORDER
        ws.row_dimensions[r].height = height


def _blank_rows(ws: Worksheet, start: int, n_rows: int, n_cols: int, height: int = 20):
    """입력하기 좋게 교차 음영 + 테두리만 깔아둔 빈 행."""
    for k in range(n_rows):
        r = start + k
        fill = ZEBRA_FILL if k % 2 else WHITE_FILL
        for i in range(1, n_cols + 1):
            c = ws.cell(row=r, column=i)
            c.fill = fill
            c.font = BODY_FONT
            c.alignment = LEFT
            c.border = BORDER
        ws.row_dimensions[r].height = height


# ── 안내 시트 ──────────────────────────────────────────────
def build_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("안내", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = SKY
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 104

    # 배너
    ws.merge_cells("A1:B1")
    ws.merge_cells("A2:B2")
    t = ws["A1"]; t.value = "부품 체결/반송 자동화 · 양산평가 입력 양식"
    t.font = BANNER_FONT; t.fill = BANNER_FILL
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s = ws["A2"]; s.value = "PRODUCTION EVALUATION — VENDOR INPUT TEMPLATE"
    s.font = BANNER_SUB; s.fill = BANNER_FILL
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 18

    sections = [
        ("작성 방법", [
            "1) 본 파일은 양식입니다. 매일 평가가 끝나면 [일일평가] 시트에 1행씩 추가하세요.",
            "2) 에러가 발생한 날에는 [에러로그] 시트에도 해당 에러 1건을 별도로 추가하세요.",
            "3) 파일명은 자유 (예: 양산평가_2026-06-17.xlsx). 한 파일에 누적 기록을 유지하세요.",
            "4) 작성 완료한 파일을 PM에게 전달하면 대시보드에 반영됩니다.",
        ]),
        ("주의사항", [
            "• 시트명(일일평가 / 에러로그)은 변경하지 마세요.",
            "• 파란 헤더 행은 그대로 유지하세요. 컬럼 이름이 바뀌면 자동 변환이 실패합니다.",
            "• 날짜는 YYYY-MM-DD (예: 2026-06-17), 시각은 h:mm (예: 14:32).",
            "• 옅은 파란색 이탤릭 '예시' 행은 참고용입니다. 실제 입력 시 덮어쓰거나 아래 빈 행에 추가하세요.",
        ]),
        ("핵심 메트릭", [
            "• 일일평가 : 그 날 수행한 총 사이클 횟수 (성공+실패 모두 포함)",
            "• 일일에러 : 그 날 발생한 에러(인시던트) 횟수",
            "• 연속성공 : 마지막 에러 이후 누적 성공 사이클 수",
        ]),
        ("에러 상세자료 (선택)", [
            "• 상세설명 : 에러 행마다 길게 적고 싶은 분석/경위. 대시보드 [＋상세] 버튼에서만 보입니다.",
            "• 사진(파일명) : 첨부 이미지 '파일명'만 쉼표로 구분해 적습니다. 예) ERR-001_1.jpg, ERR-001_2.jpg",
            "    └ 사진은 엑셀에 붙이지 말고, 파일명만 적은 뒤 이미지 파일을 엑셀과 함께(zip) 전달하세요.",
            "    └ 두 칸 모두 비워두면 [＋상세] 버튼은 표시되지 않습니다.",
        ]),
        ("문의", [
            "양식 변경·문제 발생 시 PM에게 문의하세요.",
        ]),
    ]
    row = 4
    for title, lines in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=1, value=title)
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = SECTION_BORDER
        ws.row_dimensions[row].height = 26
        row += 1
        for line in lines:
            b = ws.cell(row=row, column=2, value=line)
            b.font = NOTE_FONT
            b.alignment = LEFT
            ws.row_dimensions[row].height = 19
            row += 1
        row += 1   # 섹션 간 여백


# ── 일일평가 시트 ──────────────────────────────────────────────
def build_daily_sheet(wb: Workbook):
    ws = wb.create_sheet("일일평가")
    ws.sheet_properties.tabColor = NAVY
    cols   = ["평가일", "입실인원", "주평가내용", "일일평가", "일일에러", "연속성공", "비고"]
    widths = [15, 22, 48, 12, 12, 12, 26]
    _style_header(ws, cols, widths)

    examples = [
        ["2026-06-01", "홍길동, 김철수", "JOB 생성 - 픽업 - 적재 사이클 셋업", 42, 0, 42,  "← 예시 (덮어쓰거나 아래에 입력)"],
        ["2026-06-02", "홍길동, 김철수", "사이클 반복 안정성 검증",            78, 0, 120, "← 예시"],
    ]
    _example_rows(ws, 2, examples, len(cols), height=24)
    _blank_rows(ws, 4, 16, len(cols), height=22)


# ── 에러로그 시트 ──────────────────────────────────────────────
def build_errors_sheet(wb: Workbook):
    ws = wb.create_sheet("에러로그")
    ws.sheet_properties.tabColor = NAVY
    cols   = ["No", "발생일", "시각", "회차", "코드", "유형", "상세", "원인", "조치", "결과",
              "고객사 담당자", "업체 담당자", "상세설명", "사진(파일명)"]
    widths = [6, 14, 10, 10, 12, 18, 44, 34, 44, 13, 13, 13, 54, 26]
    _style_header(ws, cols, widths)

    examples = [
        [1, "2026-06-08", "14:32", 358, "ERR-001", "비전 인식 오류", "픽업 대상 부품의 비전 좌표 인식 실패, 로봇 정지",
         "조도 변화로 카메라 노출값 부적합 추정", "조명 LUX 재조정 + 비전 threshold 보정", "정상복귀", "김철수", "박영희",
         "현장 조도 320→210 LUX 급감 구간에서 반복. 노출 보정 후 재현 안 됨. (분석 리포트 별첨)", "ERR-001_1.jpg, ERR-001_2.jpg"],
        [2, "2026-06-17", "11:08", 953, "ERR-002", "그리퍼 그립 실패", "부품 표면 마찰계수 편차로 그리핑 실패, 자동 정지",
         "부품 표면 코팅 편차 추정", "그리퍼 압력 +5% 조정, 표면 사전검사 추가", "정상복귀", "이영희", "홍길동",
         "코팅 로트 편차로 마찰계수 0.40→0.28. 압력 상향으로 해결.", "ERR-002_grip.png"],
    ]
    _example_rows(ws, 2, examples, len(cols), height=42)
    _blank_rows(ws, 4, 14, len(cols), height=24)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_guide_sheet(wb)
    build_daily_sheet(wb)
    build_errors_sheet(wb)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[template] 생성 완료: {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
