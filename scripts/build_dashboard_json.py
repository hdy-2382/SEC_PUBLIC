"""
build_dashboard_json.py
-----------------------
업체가 보낸 .xlsx 파일을 읽어서 data/dashboard.json 으로 변환한다.

규칙:
  - data/raw/ 폴더 내에서 가장 최근에 수정된 .xlsx 파일을 입력으로 사용
  - 엑셀 시트:
      "일일평가"  → daily 배열
      "에러로그"  → errors 배열
    (시트명이 정확히 일치하지 않으면 부분 매칭으로 찾는다)
  - 출력: data/dashboard.json (UTF-8, 한글 그대로)

읽기 방식:
  - 1차: openpyxl (빠름, 일반 xlsx 전용)
  - 2차 폴백: xlwings (Excel 실행, DRM 보호된 xlsx도 처리 가능)
    → Windows + Excel 설치 필요. openpyxl이 zipfile.BadZipFile로 실패하면 자동 전환.

기대 컬럼:
  일일평가 시트  : 평가일, 입실인원, 주평가내용, 일일평가, 일일에러, 연속성공, 비고
  에러로그 시트  : No, 발생일, 시각, 회차, 코드, 유형, 상세, 원인, 조치, 결과, 고객사 담당자, 업체 담당자

'시각' 처리:
  업체가 custom h:mm 서식으로 시각을 보내면 openpyxl이 시간을 '하루의 분수'(0~1 float)로
  넘겨 0.xxxxx 형태로 깨질 수 있다. _cell_to_time() 가 이를 h:mm 으로 복원한다.
"""

from __future__ import annotations

import json
import re
import zipfile
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUT_PATH = ROOT / "data" / "dashboard.json"


# ── 컬럼 헤더 매핑 ─────────────────────────────────────────────
# 한글 헤더에서 공백/괄호 제거 후 매칭한다 (오타 방지)
DAILY_FIELD_ALIASES = {
    "date":      ["평가일", "일자", "date"],
    "personnel": ["입실인원", "입실자", "인원", "personnel"],
    "activity":  ["주평가내용", "평가내용", "내용", "activity"],
    "total":     ["일일평가", "평가횟수", "사이클", "total"],
    "errors":    ["일일에러", "에러", "errors"],
    "streak":    ["연속성공", "연속", "streak"],
    "notes":     ["비고", "메모", "notes"],
}
ERROR_FIELD_ALIASES = {
    "no":     ["no", "번호", "순번", "no."],
    "date":   ["발생일", "일자", "date"],
    "time":   ["시각", "시간", "time"],
    "cycle":  ["회차", "사이클", "cycle"],
    "code":   ["코드", "code"],
    "type":   ["유형", "타입", "type"],
    "detail": ["상세", "상세내용", "detail"],
    "cause":  ["원인", "cause"],
    "action": ["조치", "조치사항", "action"],
    "result": ["결과", "조치결과", "result"],
    # 고객사 담당자 / 업체 담당자 — 둘 다 "담당"을 포함하므로 구체적인 후보를 먼저 둔다.
    # owner_sec(고객사)를 owner(업체)보다 먼저 매핑해 교차 매칭을 방지.
    "owner_sec": ["고객사담당자", "고객사담당", "고객사", "secowner", "sec"],
    "owner":     ["업체담당자", "업체담당", "협력사담당", "업체", "협력사", "vendor", "담당", "owner"],
    # 업체가 입력하는 확장 자료 (선택) — "더 상세" 모달에서만 표시.
    #   detail_more : 긴 설명 텍스트
    #   images      : 사진 파일명(쉼표/줄바꿈 구분). 실제 파일은 data/errors/ 폴더에 둔다.
    "detail_more": ["상세설명", "추가상세", "상세자료", "추가설명", "detailmore"],
    "images":      ["사진", "이미지", "첨부파일", "첨부", "파일명", "image", "photo", "attachment"],
}

DAILY_SHEET_KEYWORDS  = ["일일평가", "일일", "daily"]
ERRORS_SHEET_KEYWORDS = ["에러로그", "에러", "error"]


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).lower()


def _find_name(names: list[str], keywords: list[str]) -> str | None:
    """시트명 리스트에서 keywords 중 하나라도 포함된 첫 이름을 반환."""
    norm_map = {n: _norm(n) for n in names}
    for kw in keywords:
        kw_n = _norm(kw)
        for name in names:
            if kw_n in norm_map[name]:
                return name
    return None


def _build_column_map(header_row, aliases: dict[str, list[str]]) -> dict[str, int]:
    """헤더 행을 보고 {필드명: 컬럼인덱스} 매핑을 만든다."""
    norm_cells = [_norm(c) for c in header_row]
    col_map: dict[str, int] = {}
    for field, candidates in aliases.items():
        for cand in candidates:
            cand_n = _norm(cand)
            for idx, cell in enumerate(norm_cells):
                if cell == cand_n or (cand_n and cand_n in cell):
                    col_map[field] = idx
                    break
            if field in col_map:
                break
    return col_map


def _find_header_row(rows: list[list], aliases: dict[str, list[str]]) -> int:
    """선두 최대 10행에서 헤더처럼 보이는 행 인덱스를 찾는다.
    vendor가 상단에 제목/병합 안내를 넣었을 때 1행이 아닐 수 있어서 자동 감지.
    헤더 후보 매칭 개수가 가장 많은 행을 선택. 매칭이 전혀 없으면 0(첫행) fallback.
    """
    best_idx, best_hits = 0, 0
    candidates_flat: list[str] = []
    for cands in aliases.values():
        candidates_flat.extend(_norm(c) for c in cands if c)
    for idx, row in enumerate(rows[:10]):
        if row is None:
            continue
        norm_cells = [_norm(c) for c in row]
        hits = 0
        for cell in norm_cells:
            if not cell:
                continue
            for cand_n in candidates_flat:
                if cell == cand_n or (cand_n and cand_n in cell):
                    hits += 1
                    break
        if hits > best_hits:
            best_idx, best_hits = idx, hits
    return best_idx


_DATE_TEXT_RE = re.compile(
    r"^\s*(\d{2,4})[\s\.\-/년]+(\d{1,2})[\s\.\-/월]+(\d{1,2})\s*일?\s*$"
)


def _try_normalize_date(s: str) -> str:
    """텍스트로 들어온 날짜를 YYYY-MM-DD 로 정규화. 매칭 안 되면 원문 반환.
    지원 포맷: 2026-06-01, 2026/06/01, 2026.6.1, 26-6-1, 2026년 6월 1일 등.
    """
    m = _DATE_TEXT_RE.match(s)
    if not m:
        return s
    y, mo, d = m.group(1), m.group(2), m.group(3)
    if len(y) == 2:
        y = f"20{y}"
    try:
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    except ValueError:
        return s


_EXCEL_EPOCH = datetime(1899, 12, 30)   # Excel 1900 leap-year 버그 보정 포함


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    # Excel 직렬 날짜 — 셀 서식이 '일반/숫자'면 float로 넘어옴.
    # 20000~80000 범위면 1954~2118년 사이 → 날짜로 해석. (bool은 int 서브클래스라 명시적으로 제외)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            f = float(v)
            if 20000 < f < 80000:
                return (_EXCEL_EPOCH + timedelta(days=f)).strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            pass
    s = str(v).strip()
    # 날짜처럼 보이면 정규화 시도 (다양한 구분자/생략 연도 등 대응)
    return _try_normalize_date(s)


def _cell_to_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


_TIME_TEXT_RE = re.compile(r"^\s*(?:(?:오전|오후|am|pm)\s*)?(\d{1,2})\s*[:시]\s*(\d{1,2})")


def _cell_to_time(v) -> str:
    """에러로그 '시각' 셀 → 'h:mm' 문자열.
    업체가 custom h:mm 서식으로 보내면 openpyxl이 시간을 '하루의 분수'(0~1 float)로
    넘겨 셀 값이 0.xxxxx 로 깨지는 문제를 보정한다. (업체 입력 방식 h:mm 에 맞춰 복원)
    """
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return f"{v.hour}:{v.minute:02d}"
    if isinstance(v, time):
        return f"{v.hour}:{v.minute:02d}"
    # 숫자 — Excel 시간 직렬값. 정수부=날짜, 소수부=하루 중 비율. 소수부만 분으로 환산.
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            frac = float(v) % 1.0
        except (OverflowError, ValueError):
            frac = None
        if frac is not None:
            total_min = round(frac * 1440) % 1440   # 분 단위 반올림 + 자정 롤오버 보정
            h, m = divmod(total_min, 60)
            return f"{h}:{m:02d}"
    # 문자열 — 이미 'HH:MM'/'오후 2:32' 등으로 들어온 경우 h:mm 로 정리해서 반환
    s = str(v).strip()
    mt = _TIME_TEXT_RE.match(s)
    if mt:
        h = int(mt.group(1))
        ampm = re.match(r"^\s*(오후|pm)", s, re.IGNORECASE)
        if ampm and h < 12:
            h += 12
        return f"{h}:{int(mt.group(2)):02d}"
    return s


def _split_images(v) -> list[str]:
    """'사진' 셀 → 파일명 리스트. 쉼표/줄바꿈/세미콜론 구분. 빈 값은 제거.
    실제 이미지 파일은 data/errors/ 폴더에 같은 이름으로 둔다(프론트가 그 경로로 로드)."""
    if v is None:
        return []
    parts = re.split(r"[,\n;]+", str(v))
    return [p.strip() for p in parts if p.strip()]


def _safe_idx(row: list, idx: int):
    """row가 짧으면 None 반환 — vendor가 컬럼 수를 다르게 보낼 때 IndexError 방지."""
    return row[idx] if idx is not None and 0 <= idx < len(row) else None


def _parse_daily(rows: list[list]) -> list[dict]:
    if not rows:
        return []
    header_idx = _find_header_row(rows, DAILY_FIELD_ALIASES)
    header = rows[header_idx]
    body = rows[header_idx + 1:]
    cmap = _build_column_map(header, DAILY_FIELD_ALIASES)
    if "date" not in cmap or "total" not in cmap:
        raise SystemExit(
            f"[일일평가] 시트에서 필수 컬럼(평가일/일일평가)을 찾지 못했습니다. "
            f"감지된 헤더(행 {header_idx + 1}): {[str(h) for h in header]}"
        )

    out = []
    for row in body:
        if row is None or all(c is None or c == "" for c in row):
            continue
        date_val = _cell_to_str(_safe_idx(row, cmap["date"]))
        if not date_val:
            continue
        # YYYY-MM-DD 형식이 아니면 skip (헤더 잔여 행이나 잘못된 행 한 번 더 거름)
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_val):
            continue
        out.append({
            "date":      date_val,
            "personnel": _cell_to_str(_safe_idx(row, cmap.get("personnel"))),
            "activity":  _cell_to_str(_safe_idx(row, cmap.get("activity"))),
            "total":     max(0, _cell_to_int(_safe_idx(row, cmap["total"]))),
            "errors":    max(0, _cell_to_int(_safe_idx(row, cmap.get("errors")))),
            "streak":    max(0, _cell_to_int(_safe_idx(row, cmap.get("streak")))),
            "notes":     _cell_to_str(_safe_idx(row, cmap.get("notes"))),
        })
    out.sort(key=lambda r: r["date"])
    return out


def _parse_errors(rows: list[list]) -> list[dict]:
    if not rows:
        return []
    header_idx = _find_header_row(rows, ERROR_FIELD_ALIASES)
    header = rows[header_idx]
    body = rows[header_idx + 1:]
    cmap = _build_column_map(header, ERROR_FIELD_ALIASES)
    out = []
    for row in body:
        if row is None or all(c is None or c == "" for c in row):
            continue
        no_val = _cell_to_int(_safe_idx(row, cmap.get("no")))
        date_val = _cell_to_str(_safe_idx(row, cmap.get("date")))
        if not no_val and not date_val:
            continue
        out.append({
            "no":        no_val,
            "date":      date_val,
            "time":      _cell_to_time(_safe_idx(row, cmap.get("time"))),
            "cycle":     _cell_to_int(_safe_idx(row, cmap.get("cycle"))),
            "code":      _cell_to_str(_safe_idx(row, cmap.get("code"))),
            "type":      _cell_to_str(_safe_idx(row, cmap.get("type"))),
            "detail":    _cell_to_str(_safe_idx(row, cmap.get("detail"))),
            "cause":     _cell_to_str(_safe_idx(row, cmap.get("cause"))),
            "action":    _cell_to_str(_safe_idx(row, cmap.get("action"))),
            "result":    _cell_to_str(_safe_idx(row, cmap.get("result"))),
            "owner_sec": _cell_to_str(_safe_idx(row, cmap.get("owner_sec"))),
            "owner":     _cell_to_str(_safe_idx(row, cmap.get("owner"))),
            # 확장 자료(선택): 더 상세 모달에서만 사용
            "detailMore": _cell_to_str(_safe_idx(row, cmap.get("detail_more"))),
            "images":     _split_images(_safe_idx(row, cmap.get("images"))),
        })
    out.sort(key=lambda r: r.get("no", 0))
    return out


def _pick_latest_xlsx() -> Path:
    xlsxs = sorted(
        [p for p in RAW_DIR.glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not xlsxs:
        raise SystemExit(f"data/raw/ 폴더에 .xlsx 파일이 없습니다. ({RAW_DIR})")
    if len(xlsxs) > 1:
        print(f"[build] ⚠ data/raw/ 에 xlsx 파일이 {len(xlsxs)}개 있습니다. 최신(mtime) 파일을 사용합니다:")
        for p in xlsxs:
            ts = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
            marker = "→" if p == xlsxs[0] else " "
            print(f"        {marker} {p.name}  (mtime {ts})")
        print(f"        다른 파일을 쓰려면 불필요한 파일을 제거하거나 사용할 파일을 다시 저장(터치)하세요.")
    return xlsxs[0]


# ── 시트 로딩: openpyxl 우선, 실패하면 xlwings ────────────────────
def _load_via_openpyxl(src: Path) -> tuple[list[list], list[list], list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(src, data_only=True)
    names = wb.sheetnames
    daily_name  = _find_name(names, DAILY_SHEET_KEYWORDS)
    errors_name = _find_name(names, ERRORS_SHEET_KEYWORDS)

    if daily_name is None:
        raise SystemExit(f"'일일평가' 시트를 찾지 못했습니다. 시트 목록: {names}")

    daily_rows  = [list(r) for r in wb[daily_name].iter_rows(values_only=True)]
    errors_rows = (
        [list(r) for r in wb[errors_name].iter_rows(values_only=True)]
        if errors_name else []
    )
    return daily_rows, errors_rows, names


def _xlwings_sheet_rows(sheet) -> list[list]:
    rng = sheet.used_range
    val = rng.value
    if val is None:
        return []
    if not isinstance(val, list):
        return [[val]]
    if val and not isinstance(val[0], list):
        # 1D 결과 — 단일 행 or 단일 열
        if rng.rows.count == 1:
            return [val]
        return [[v] for v in val]
    return val


def _load_via_xlwings(src: Path) -> tuple[list[list], list[list], list[str]]:
    try:
        import xlwings as xw
    except ImportError as e:
        raise SystemExit(
            "xlwings 미설치. DRM 보호 파일을 읽으려면 'pip install xlwings' 후 재시도. "
            "(Windows + Excel 설치 필수)"
        ) from e

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    try:
        wb = app.books.open(str(src), update_links=False, read_only=True)
        try:
            names = [s.name for s in wb.sheets]
            daily_name  = _find_name(names, DAILY_SHEET_KEYWORDS)
            errors_name = _find_name(names, ERRORS_SHEET_KEYWORDS)

            if daily_name is None:
                raise SystemExit(f"'일일평가' 시트를 찾지 못했습니다. 시트 목록: {names}")

            daily_rows  = _xlwings_sheet_rows(wb.sheets[daily_name])
            errors_rows = (
                _xlwings_sheet_rows(wb.sheets[errors_name])
                if errors_name else []
            )
            return daily_rows, errors_rows, names
        finally:
            wb.close()
    finally:
        app.quit()


def _load_workbook_rows(src: Path) -> tuple[list[list], list[list]]:
    try:
        daily_rows, errors_rows, _ = _load_via_openpyxl(src)
        return daily_rows, errors_rows
    except zipfile.BadZipFile:
        # DRM 래핑 추정 — openpyxl은 zip 구조가 아니라고 거부함
        print("[build] openpyxl 실패 (DRM 추정) → xlwings로 Excel 통한 재시도")
        daily_rows, errors_rows, _ = _load_via_xlwings(src)
        return daily_rows, errors_rows


def main():
    src = _pick_latest_xlsx()
    print(f"[build] 입력 파일: {src.name}")

    daily_rows, errors_rows = _load_workbook_rows(src)
    daily  = _parse_daily(daily_rows)
    errors = _parse_errors(errors_rows)

    out = {
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source":      src.name,
        "daily":       daily,
        "errors":      errors,
    }
    OUT_PATH.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[build] 출력: {OUT_PATH.relative_to(ROOT)}  (daily {len(daily)}건, errors {len(errors)}건)")


if __name__ == "__main__":
    main()
