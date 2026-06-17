"""
generate_sample_submission.py
-----------------------------
업체가 "이대로 따라서 보내면 되는" 제출 샘플을 생성한다.

생성물:
  samples/제출샘플_양산평가.xlsx   ← 채워진 예시 (안내 + 일일평가 + 에러로그)
  samples/photos/ERR-001_1.jpg     ← 에러로그 '사진(파일명)' 과 이름이 일치하는 샘플 사진
  samples/photos/ERR-001_2.jpg
  samples/photos/ERR-002_grip.png

요지:
  - 엑셀에는 사진을 붙이지 않고 '사진(파일명)' 칸에 파일명만 적는다.
  - 사진 파일들은 엑셀과 함께(예: zip) 전달한다. → PM이 data/errors/ 에 같은 이름으로 넣는다.

build_dashboard_json.py 가 읽는 시트명·컬럼명을 그대로 사용한다.
"""

from __future__ import annotations

from datetime import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "samples"
PHOTO_DIR = OUT_DIR / "photos"
XLSX_PATH = OUT_DIR / "제출샘플_양산평가.xlsx"

# ── 엑셀 스타일 (vendor_template 과 톤 통일) ─────────────────
NAVY_DEEP = "0C1B36"
LINE = "D6D2C4"
BG_ALT = "FAF8F2"

HEADER_FILL = PatternFill("solid", fgColor=NAVY_DEEP)
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color=NAVY_DEEP)
SECTION_FONT = Font(name="맑은 고딕", size=12, bold=True, color="1A2942")
NOTE_FONT = Font(name="맑은 고딕", size=10, color="3D4147")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header(ws, cols, widths):
    for i, (name, width) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _row(ws, r, values, *, time_cols=()):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = LEFT
        c.border = BORDER
        if i in time_cols:
            c.number_format = "h:mm"   # 업체 실제 입력 방식(custom h:mm) 재현
            c.alignment = CENTER
    ws.row_dimensions[r].height = 30


# ── 안내 시트 ──────────────────────────────────────────────
def build_guide(wb: Workbook):
    ws = wb.create_sheet("안내", 0)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 88
    ws["A1"] = "양산평가 제출 샘플 — 이 형식 그대로 작성해 주세요"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    rows = [
        ("", ""),
        ("작성 방법", ""),
        ("", "1) [일일평가] 시트 — 매일 평가가 끝나면 1행씩 추가합니다."),
        ("", "2) [에러로그] 시트 — 에러가 발생한 날마다 1행씩 추가합니다."),
        ("", "3) 시트명/헤더(파란 행)는 바꾸지 마세요. 순서가 달라도 헤더 이름으로 자동 인식됩니다."),
        ("", "4) 날짜 YYYY-MM-DD, 시각 h:mm (예: 14:32) — 본 샘플의 셀 서식 그대로 두면 됩니다."),
        ("", ""),
        ("에러 상세자료 (선택)", ""),
        ("", "• 상세설명 : 길게 적고 싶은 분석/경위. 대시보드의 [＋상세] 버튼에서만 보입니다."),
        ("", "• 사진(파일명) : 첨부할 이미지 '파일명'만 쉼표로 구분해 적습니다. 예) ERR-001_1.jpg, ERR-001_2.jpg"),
        ("", "      └ 엑셀에 사진을 '붙여넣지' 마세요. 파일명만 적고, 사진 파일은 따로 보냅니다."),
        ("", "      └ 사진 파일명 규칙(권장): 에러코드_순번.확장자  예) ERR-001_1.jpg / ERR-002_grip.png"),
        ("", ""),
        ("전달 방법", ""),
        ("", "• 이 엑셀 1개 + 사진 파일들을 함께(zip 권장) PM에게 전달하세요."),
        ("", "• 사진 폴더 구성은 본 샘플의 samples/photos/ 를 참고하세요. (파일명이 엑셀과 일치해야 함)"),
        ("", ""),
        ("문의", ""),
        ("", "양식·작성 문의는 PM에게 연락 주세요."),
    ]
    for off, (lbl, body) in enumerate(rows, start=2):
        a = ws.cell(row=off, column=1, value=lbl)
        b = ws.cell(row=off, column=2, value=body)
        a.font = SECTION_FONT if lbl and lbl[-1] not in ")" and lbl in (
            "작성 방법", "에러 상세자료 (선택)", "전달 방법", "문의") else BODY_FONT
        b.font = NOTE_FONT
        b.alignment = LEFT


# ── 일일평가 시트 (채워진 예시) ───────────────────────────────
def build_daily(wb: Workbook):
    ws = wb.create_sheet("일일평가")
    cols = ["평가일", "입실인원", "주평가내용", "일일평가", "일일에러", "연속성공", "비고"]
    widths = [14, 20, 46, 12, 12, 12, 28]
    _header(ws, cols, widths)
    data = [
        ["2026-06-01", "김철수, 이영희", "JOB 생성 - 픽업 - 적재 사이클 셋업", 42, 0, 42, "초기 셋업, 안정"],
        ["2026-06-02", "김철수, 이영희", "사이클 반복 안정성 검증", 78, 0, 120, "정상"],
        ["2026-06-03", "김철수, 이영희", "연속 적재 정밀도 점검", 95, 0, 215, "정상"],
        ["2026-06-04", "김철수, 이영희", "비전 인식 파라미터 튜닝", 60, 1, 0, "ERR-001 발생 → 연속 리셋"],
        ["2026-06-05", "김철수, 이영희", "노출 보정 후 재검증", 88, 0, 88, "정상 복귀"],
        ["2026-06-06", "김철수, 이영희", "그리퍼 압력 캘리브레이션", 50, 1, 0, "ERR-002 발생"],
        ["2026-06-07", "김철수, 이영희", "표면 검사 단계 추가 검증", 102, 0, 102, "정상"],
    ]
    for i, d in enumerate(data, start=2):
        _row(ws, i, d)


# ── 에러로그 시트 (상세설명 + 사진 파일명 채워진 예시) ──────────
def build_errors(wb: Workbook):
    ws = wb.create_sheet("에러로그")
    cols = ["No", "발생일", "시각", "회차", "코드", "유형", "상세", "원인", "조치", "결과",
            "고객사 담당자", "업체 담당자", "상세설명", "사진(파일명)"]
    widths = [6, 14, 10, 10, 12, 18, 42, 34, 42, 14, 14, 14, 56, 28]
    _header(ws, cols, widths)
    data = [
        [1, "2026-06-04", time(14, 32), 358, "ERR-001", "비전 인식 오류",
         "픽업 대상 부품의 비전 좌표 인식 실패, 로봇 정지",
         "조도 변화로 카메라 노출값 부적합 추정",
         "조명 LUX 재조정 + 비전 threshold 보정", "정상복귀", "김철수", "박영희",
         "현장 조도 320→210 LUX 급감 구간에서 반복 발생. 노출 EV+0.7 보정 후 재현 안 됨. "
         "비전 threshold 0.82→0.74 하향으로 픽업 성공률 회복.",
         "ERR-001_1.jpg, ERR-001_2.jpg"],
        [2, "2026-06-06", time(11, 8), 953, "ERR-002", "그리퍼 그립 실패",
         "부품 표면 마찰계수 편차로 그리핑 실패, 자동 정지",
         "부품 표면 코팅 로트 편차 추정",
         "그리퍼 압력 +5% 조정 + 표면 사전검사 추가", "정상복귀", "이영희", "홍길동",
         "코팅 로트 편차로 표면 마찰계수 0.40→0.28. 압력 상향으로 해결. 후속 로트 입고검사 강화 요청.",
         "ERR-002_grip.png"],
    ]
    for i, d in enumerate(data, start=2):
        _row(ws, i, d, time_cols=(3,))
        ws.row_dimensions[i].height = 46


# ── 샘플 사진 생성 (PIL) ──────────────────────────────────────
def _font(size: int):
    for name in ("DejaVuSans-Bold.ttf", "DejaVuSans.ttf"):
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def _make_photo(path: Path, label: str, caption: str, bg: tuple[int, int, int]):
    W, H = 800, 600
    img = Image.new("RGB", (W, H), bg)
    d = ImageDraw.Draw(img)
    # 테두리 + 대각선(샘플 워터마크 느낌)
    d.rectangle([16, 16, W - 16, H - 16], outline=(255, 255, 255), width=4)
    d.line([16, 16, W - 16, H - 16], fill=(255, 255, 255), width=1)
    d.line([16, H - 16, W - 16, 16], fill=(255, 255, 255), width=1)
    # 중앙 라벨
    f_big, f_mid, f_small = _font(54), _font(30), _font(22)

    def centered(txt, font, y, fill=(255, 255, 255)):
        l, t, r, b = d.textbbox((0, 0), txt, font=font)
        d.text(((W - (r - l)) / 2, y), txt, font=font, fill=fill)

    centered("SAMPLE", f_big, 150)
    centered("error site photo", f_mid, 240, (220, 230, 245))
    centered(label, f_mid, 360)
    centered(caption, f_small, 430, (210, 220, 235))
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() in (".jpg", ".jpeg"):
        img.save(path, "JPEG", quality=85)
    else:
        img.save(path, "PNG")


def build_photos():
    _make_photo(PHOTO_DIR / "ERR-001_1.jpg", "ERR-001_1.jpg",
                "vision misdetect - before fix", (30, 74, 122))
    _make_photo(PHOTO_DIR / "ERR-001_2.jpg", "ERR-001_2.jpg",
                "after exposure / threshold tuning", (47, 93, 63))
    _make_photo(PHOTO_DIR / "ERR-002_grip.png", "ERR-002_grip.png",
                "gripper slip - surface coating", (139, 46, 31))


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_guide(wb)
    build_daily(wb)
    build_errors(wb)
    wb.save(XLSX_PATH)
    build_photos()
    print(f"[sample] 엑셀: {XLSX_PATH.relative_to(ROOT)}")
    for p in sorted(PHOTO_DIR.glob('*')):
        print(f"[sample] 사진: {p.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
